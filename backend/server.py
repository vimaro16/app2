from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from enum import Enum
import resend
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'suerteapp-secret-key-change-in-production')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Resend Email Config
resend.api_key = os.environ.get('RESEND_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', '')

# Create the main app
app = FastAPI(title="SuerteApp API", version="1.0.0")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============ ENUMS ============
class UserRole(str, Enum):
    ADMIN = "admin"
    EDITOR = "editor"
    USER = "user"

class RaffleStatus(str, Enum):
    ACTIVE = "active"
    INACTIVE = "inactive"
    FINISHED = "finished"

class SlotStatus(str, Enum):
    AVAILABLE = "available"
    PENDING = "pending"
    PAID = "paid"

class PaymentStatus(str, Enum):
    PENDING = "pending"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"

class PaymentMethod(str, Enum):
    STRIPE = "stripe"
    PAYPAL = "paypal"
    TRANSFER = "transfer"
    WHATSAPP = "whatsapp"

# Sponsor Commission Rate
SPONSOR_COMMISSION_RATE = 0.10  # 10%

def generate_sponsor_code():
    """Generate a unique 8-character sponsor code"""
    import random
    import string
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))

# ============ MODELS ============
class UserCreate(BaseModel):
    full_name: str = Field(..., min_length=2)
    cedula: str = Field(..., min_length=5)
    whatsapp: str = Field(..., min_length=10)
    email: str
    password: str = Field(..., min_length=6)
    sponsor_code: Optional[str] = None  # Code used during registration

class UserLogin(BaseModel):
    email: str
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    full_name: str
    cedula: str
    whatsapp: str
    email: str
    role: UserRole
    sponsor_code: str
    referred_by: Optional[str] = None
    created_at: str

class UserRoleUpdate(BaseModel):
    role: UserRole

class PaymentMethodConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    method_type: str
    name: str
    is_active: bool
    details: Dict[str, Any]
    created_at: str

class PaymentMethodCreate(BaseModel):
    method_type: str
    name: str
    is_active: bool = True
    details: Dict[str, Any] = {}

class RaffleCreate(BaseModel):
    title: str = Field(..., min_length=3)
    description: str
    image_url: Optional[str] = None
    video_url: Optional[str] = None
    slot_price: float = Field(..., gt=0)

class RaffleUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    image_url: Optional[str] = None
    video_url: Optional[str] = None
    slot_price: Optional[float] = None
    status: Optional[RaffleStatus] = None
    finish_text: Optional[str] = None
    finish_image_url: Optional[str] = None
    finish_video_url: Optional[str] = None

class RaffleResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    title: str
    description: str
    image_url: Optional[str]
    video_url: Optional[str]
    slot_price: float
    status: RaffleStatus
    created_by: str
    created_at: str
    finished_at: Optional[str]
    finish_text: Optional[str]
    finish_image_url: Optional[str]
    finish_video_url: Optional[str]

class SlotResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    number: int
    status: SlotStatus
    initials: Optional[str]
    user_id: Optional[str]

class SlotPurchaseRequest(BaseModel):
    slot_numbers: List[int] = Field(..., min_length=1)
    payment_method: PaymentMethod
    sponsor_code: Optional[str] = None  # Sponsor code for commission

class PaymentTransactionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    raffle_id: str
    slot_numbers: List[int]
    amount: float
    payment_method: str
    status: PaymentStatus
    session_id: Optional[str]
    sponsor_code: Optional[str] = None
    sponsor_commission: Optional[float] = None
    created_at: str

class SponsorEarningResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    sponsor_id: str
    transaction_id: str
    raffle_id: str
    buyer_name: str
    slot_numbers: List[int]
    commission_amount: float
    status: str  # pending, paid
    created_at: str

class PaymentRequestCreate(BaseModel):
    bank_name: str = Field(..., min_length=2)
    account_number: str = Field(..., min_length=5)
    account_holder: str = Field(..., min_length=2)
    amount: float = Field(..., gt=0)

class PaymentRequestResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    sponsor_id: str
    sponsor_name: str
    sponsor_email: str
    sponsor_whatsapp: str
    bank_name: str
    account_number: str
    account_holder: str
    amount: float
    status: str  # pending, approved, rejected, paid
    created_at: str
    processed_at: Optional[str] = None

class WeeklySummary(BaseModel):
    week: str
    total_earnings: float
    pending_earnings: float
    paid_earnings: float
    sales_count: int
    can_request_payment: bool

# ============ AUTHENTICATION ============
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

async def require_admin(user: dict = Depends(get_current_user)):
    if user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Se requieren permisos de administrador")
    return user

async def require_editor_or_admin(user: dict = Depends(get_current_user)):
    if user["role"] not in [UserRole.ADMIN, UserRole.EDITOR]:
        raise HTTPException(status_code=403, detail="Se requieren permisos de editor o administrador")
    return user

# ============ AUTH ROUTES ============
@api_router.post("/auth/register", response_model=dict)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"$or": [{"email": user_data.email}, {"cedula": user_data.cedula}]})
    if existing:
        raise HTTPException(status_code=400, detail="Email o cédula ya registrados")
    
    # Check if this is the first user (make admin)
    user_count = await db.users.count_documents({})
    role = UserRole.ADMIN if user_count == 0 else UserRole.USER
    
    # Generate unique sponsor code
    sponsor_code = generate_sponsor_code()
    while await db.users.find_one({"sponsor_code": sponsor_code}):
        sponsor_code = generate_sponsor_code()
    
    # Check if referred by someone
    referred_by = None
    if user_data.sponsor_code:
        referrer = await db.users.find_one({"sponsor_code": user_data.sponsor_code}, {"_id": 0})
        if referrer:
            referred_by = referrer["id"]
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "full_name": user_data.full_name,
        "cedula": user_data.cedula,
        "whatsapp": user_data.whatsapp,
        "email": user_data.email,
        "password": hash_password(user_data.password),
        "role": role,
        "sponsor_code": sponsor_code,
        "referred_by": referred_by,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    token = create_token(user_id, role)
    return {"token": token, "user": {k: v for k, v in user_doc.items() if k not in ["password", "_id"]}}

@api_router.post("/auth/login", response_model=dict)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    token = create_token(user["id"], user["role"])
    return {"token": token, "user": {k: v for k, v in user.items() if k != "password"}}

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return user

# ============ USER MANAGEMENT ROUTES ============
@api_router.get("/users", response_model=List[UserResponse])
async def get_users(admin: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return users

@api_router.put("/users/{user_id}/role", response_model=UserResponse)
async def update_user_role(user_id: str, role_data: UserRoleUpdate, admin: dict = Depends(require_admin)):
    result = await db.users.find_one_and_update(
        {"id": user_id},
        {"$set": {"role": role_data.role}},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {k: v for k, v in result.items() if k not in ["password", "_id"]}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin)):
    if admin["id"] == user_id:
        raise HTTPException(status_code=400, detail="No puedes eliminarte a ti mismo")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"message": "Usuario eliminado"}

# ============ PAYMENT METHODS ROUTES ============
@api_router.get("/payment-methods", response_model=List[PaymentMethodConfig])
async def get_payment_methods():
    methods = await db.payment_methods.find({}, {"_id": 0}).to_list(100)
    return methods

@api_router.post("/payment-methods", response_model=PaymentMethodConfig)
async def create_payment_method(method_data: PaymentMethodCreate, admin: dict = Depends(require_admin)):
    method_doc = {
        "id": str(uuid.uuid4()),
        "method_type": method_data.method_type,
        "name": method_data.name,
        "is_active": method_data.is_active,
        "details": method_data.details,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payment_methods.insert_one(method_doc)
    return {k: v for k, v in method_doc.items() if k != "_id"}

@api_router.put("/payment-methods/{method_id}", response_model=PaymentMethodConfig)
async def update_payment_method(method_id: str, method_data: PaymentMethodCreate, admin: dict = Depends(require_admin)):
    result = await db.payment_methods.find_one_and_update(
        {"id": method_id},
        {"$set": {
            "method_type": method_data.method_type,
            "name": method_data.name,
            "is_active": method_data.is_active,
            "details": method_data.details
        }},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Método de pago no encontrado")
    return {k: v for k, v in result.items() if k != "_id"}

@api_router.delete("/payment-methods/{method_id}")
async def delete_payment_method(method_id: str, admin: dict = Depends(require_admin)):
    result = await db.payment_methods.delete_one({"id": method_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Método de pago no encontrado")
    return {"message": "Método de pago eliminado"}

# ============ RAFFLE ROUTES ============
@api_router.post("/raffles", response_model=RaffleResponse)
async def create_raffle(raffle_data: RaffleCreate, user: dict = Depends(require_editor_or_admin)):
    raffle_id = str(uuid.uuid4())
    raffle_doc = {
        "id": raffle_id,
        "title": raffle_data.title,
        "description": raffle_data.description,
        "image_url": raffle_data.image_url,
        "video_url": raffle_data.video_url,
        "slot_price": raffle_data.slot_price,
        "status": RaffleStatus.ACTIVE,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "finished_at": None,
        "finish_text": None,
        "finish_image_url": None,
        "finish_video_url": None
    }
    await db.raffles.insert_one(raffle_doc)
    
    # Create 100 slots
    slots = [
        {"raffle_id": raffle_id, "number": i, "status": SlotStatus.AVAILABLE, "initials": None, "user_id": None}
        for i in range(1, 101)
    ]
    await db.slots.insert_many(slots)
    
    return {k: v for k, v in raffle_doc.items() if k != "_id"}

@api_router.get("/raffles", response_model=List[RaffleResponse])
async def get_raffles(status: Optional[RaffleStatus] = None):
    query = {}
    if status:
        query["status"] = status
    raffles = await db.raffles.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return raffles

@api_router.get("/raffles/active", response_model=List[RaffleResponse])
async def get_active_raffles():
    raffles = await db.raffles.find({"status": RaffleStatus.ACTIVE}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return raffles

@api_router.get("/raffles/{raffle_id}", response_model=RaffleResponse)
async def get_raffle(raffle_id: str):
    raffle = await db.raffles.find_one({"id": raffle_id}, {"_id": 0})
    if not raffle:
        raise HTTPException(status_code=404, detail="Rifa no encontrada")
    return raffle

@api_router.put("/raffles/{raffle_id}", response_model=RaffleResponse)
async def update_raffle(raffle_id: str, raffle_data: RaffleUpdate, user: dict = Depends(require_editor_or_admin)):
    update_dict = {k: v for k, v in raffle_data.model_dump().items() if v is not None}
    
    # If deactivating/finishing, set finished_at
    if raffle_data.status in [RaffleStatus.INACTIVE, RaffleStatus.FINISHED]:
        update_dict["finished_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.raffles.find_one_and_update(
        {"id": raffle_id},
        {"$set": update_dict},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Rifa no encontrada")
    return {k: v for k, v in result.items() if k != "_id"}

@api_router.delete("/raffles/{raffle_id}")
async def delete_raffle(raffle_id: str, admin: dict = Depends(require_admin)):
    result = await db.raffles.delete_one({"id": raffle_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rifa no encontrada")
    await db.slots.delete_many({"raffle_id": raffle_id})
    return {"message": "Rifa eliminada"}

# ============ SLOTS ROUTES ============
@api_router.get("/raffles/{raffle_id}/slots", response_model=List[SlotResponse])
async def get_raffle_slots(raffle_id: str):
    raffle = await db.raffles.find_one({"id": raffle_id})
    if not raffle:
        raise HTTPException(status_code=404, detail="Rifa no encontrada")
    slots = await db.slots.find({"raffle_id": raffle_id}, {"_id": 0, "raffle_id": 0}).sort("number", 1).to_list(100)
    return slots

@api_router.post("/raffles/{raffle_id}/slots/reserve")
async def reserve_slots(raffle_id: str, request: SlotPurchaseRequest, user: dict = Depends(get_current_user)):
    raffle = await db.raffles.find_one({"id": raffle_id}, {"_id": 0})
    if not raffle:
        raise HTTPException(status_code=404, detail="Rifa no encontrada")
    if raffle["status"] != RaffleStatus.ACTIVE:
        raise HTTPException(status_code=400, detail="Esta rifa no está activa")
    
    # Check if slots are available
    for slot_num in request.slot_numbers:
        slot = await db.slots.find_one({"raffle_id": raffle_id, "number": slot_num})
        if not slot or slot["status"] != SlotStatus.AVAILABLE:
            raise HTTPException(status_code=400, detail=f"La casilla {slot_num} no está disponible")
    
    # Get user initials
    name_parts = user["full_name"].split()
    initials = "".join([p[0].upper() for p in name_parts[:2]]) if len(name_parts) > 1 else user["full_name"][:2].upper()
    
    # Reserve slots
    await db.slots.update_many(
        {"raffle_id": raffle_id, "number": {"$in": request.slot_numbers}},
        {"$set": {"status": SlotStatus.PENDING, "user_id": user["id"], "initials": initials}}
    )
    
    amount = len(request.slot_numbers) * raffle["slot_price"]
    
    # Check sponsor code and calculate commission
    sponsor_id = None
    sponsor_commission = 0
    if request.sponsor_code:
        sponsor = await db.users.find_one({"sponsor_code": request.sponsor_code}, {"_id": 0})
        if sponsor and sponsor["id"] != user["id"]:  # Can't use own code
            sponsor_id = sponsor["id"]
            sponsor_commission = float(amount * SPONSOR_COMMISSION_RATE)
    
    # Create payment transaction
    transaction_id = str(uuid.uuid4())
    transaction = {
        "id": transaction_id,
        "user_id": user["id"],
        "raffle_id": raffle_id,
        "slot_numbers": request.slot_numbers,
        "amount": float(amount),
        "payment_method": request.payment_method,
        "status": PaymentStatus.PENDING,
        "session_id": None,
        "sponsor_code": request.sponsor_code,
        "sponsor_id": sponsor_id,
        "sponsor_commission": sponsor_commission,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payment_transactions.insert_one(transaction)
    
    return {"transaction_id": transaction_id, "amount": amount, "slots": request.slot_numbers}

# ============ PAYMENT ROUTES ============
@api_router.post("/payments/stripe/create-session")
async def create_stripe_session(request: Request, transaction_id: str, user: dict = Depends(get_current_user)):
    transaction = await db.payment_transactions.find_one({"id": transaction_id, "user_id": user["id"]}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout, CheckoutSessionRequest
        
        api_key = os.environ.get('STRIPE_API_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="Stripe no configurado")
        
        body = await request.json()
        origin_url = body.get("origin_url", str(request.base_url))
        
        webhook_url = f"{str(request.base_url)}api/webhook/stripe"
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url=webhook_url)
        
        success_url = f"{origin_url}/payment/success?session_id={{CHECKOUT_SESSION_ID}}"
        cancel_url = f"{origin_url}/raffle/{transaction['raffle_id']}"
        
        checkout_request = CheckoutSessionRequest(
            amount=float(transaction["amount"]),
            currency="usd",
            success_url=success_url,
            cancel_url=cancel_url,
            metadata={"transaction_id": transaction_id, "user_id": user["id"]}
        )
        
        session = await stripe_checkout.create_checkout_session(checkout_request)
        
        await db.payment_transactions.update_one(
            {"id": transaction_id},
            {"$set": {"session_id": session.session_id, "status": PaymentStatus.PENDING}}
        )
        
        return {"url": session.url, "session_id": session.session_id}
    except Exception as e:
        logger.error(f"Stripe error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/payments/stripe/status/{session_id}")
async def get_stripe_status(session_id: str, user: dict = Depends(get_current_user)):
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout
        
        api_key = os.environ.get('STRIPE_API_KEY')
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url="")
        
        status = await stripe_checkout.get_checkout_status(session_id)
        
        if status.payment_status == "paid":
            transaction = await db.payment_transactions.find_one({"session_id": session_id}, {"_id": 0})
            if transaction and transaction["status"] != PaymentStatus.COMPLETED:
                await db.payment_transactions.update_one(
                    {"session_id": session_id},
                    {"$set": {"status": PaymentStatus.COMPLETED}}
                )
                await db.slots.update_many(
                    {"raffle_id": transaction["raffle_id"], "number": {"$in": transaction["slot_numbers"]}},
                    {"$set": {"status": SlotStatus.PAID}}
                )
                # Register sponsor commission
                if transaction.get("sponsor_id") and transaction.get("sponsor_commission", 0) > 0:
                    buyer = await db.users.find_one({"id": transaction["user_id"]}, {"_id": 0})
                    earning_doc = {
                        "id": str(uuid.uuid4()),
                        "sponsor_id": transaction["sponsor_id"],
                        "transaction_id": transaction["id"],
                        "raffle_id": transaction["raffle_id"],
                        "buyer_id": transaction["user_id"],
                        "buyer_name": buyer["full_name"] if buyer else "Usuario",
                        "slot_numbers": transaction["slot_numbers"],
                        "commission_amount": transaction["sponsor_commission"],
                        "status": "pending",
                        "week_year": datetime.now(timezone.utc).strftime("%Y-W%W"),
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.sponsor_earnings.insert_one(earning_doc)
        
        return {"status": status.status, "payment_status": status.payment_status}
    except Exception as e:
        logger.error(f"Stripe status error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout
        
        api_key = os.environ.get('STRIPE_API_KEY')
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url="")
        
        body = await request.body()
        signature = request.headers.get("Stripe-Signature")
        
        webhook_response = await stripe_checkout.handle_webhook(body, signature)
        
        if webhook_response.payment_status == "paid":
            transaction = await db.payment_transactions.find_one({"session_id": webhook_response.session_id}, {"_id": 0})
            if transaction:
                await db.payment_transactions.update_one(
                    {"session_id": webhook_response.session_id},
                    {"$set": {"status": PaymentStatus.COMPLETED}}
                )
                await db.slots.update_many(
                    {"raffle_id": transaction["raffle_id"], "number": {"$in": transaction["slot_numbers"]}},
                    {"$set": {"status": SlotStatus.PAID}}
                )
        
        return {"status": "processed"}
    except Exception as e:
        logger.error(f"Stripe webhook error: {e}")
        return {"status": "error"}

@api_router.post("/payments/transfer/confirm")
async def confirm_transfer(transaction_id: str, admin: dict = Depends(require_admin)):
    transaction = await db.payment_transactions.find_one({"id": transaction_id}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transacción no encontrada")
    
    await db.payment_transactions.update_one(
        {"id": transaction_id},
        {"$set": {"status": PaymentStatus.COMPLETED}}
    )
    
    await db.slots.update_many(
        {"raffle_id": transaction["raffle_id"], "number": {"$in": transaction["slot_numbers"]}},
        {"$set": {"status": SlotStatus.PAID}}
    )
    
    # Register sponsor commission if applicable
    if transaction.get("sponsor_id") and transaction.get("sponsor_commission", 0) > 0:
        buyer = await db.users.find_one({"id": transaction["user_id"]}, {"_id": 0})
        earning_doc = {
            "id": str(uuid.uuid4()),
            "sponsor_id": transaction["sponsor_id"],
            "transaction_id": transaction_id,
            "raffle_id": transaction["raffle_id"],
            "buyer_id": transaction["user_id"],
            "buyer_name": buyer["full_name"] if buyer else "Usuario",
            "slot_numbers": transaction["slot_numbers"],
            "commission_amount": transaction["sponsor_commission"],
            "status": "pending",
            "week_year": datetime.now(timezone.utc).strftime("%Y-W%W"),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.sponsor_earnings.insert_one(earning_doc)
    
    return {"message": "Pago confirmado"}

@api_router.get("/payments/pending", response_model=List[PaymentTransactionResponse])
async def get_pending_payments(admin: dict = Depends(require_admin)):
    transactions = await db.payment_transactions.find(
        {"status": PaymentStatus.PENDING},
        {"_id": 0}
    ).to_list(100)
    return transactions

@api_router.get("/payments/my-transactions", response_model=List[PaymentTransactionResponse])
async def get_my_transactions(user: dict = Depends(get_current_user)):
    transactions = await db.payment_transactions.find(
        {"user_id": user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return transactions

# ============ BLOG ROUTES (ARCHIVED RAFFLES) ============
@api_router.get("/blog", response_model=List[RaffleResponse])
async def get_archived_raffles():
    raffles = await db.raffles.find(
        {"status": {"$in": [RaffleStatus.INACTIVE, RaffleStatus.FINISHED]}},
        {"_id": 0}
    ).sort("finished_at", -1).to_list(100)
    return raffles

@api_router.get("/blog/{raffle_id}", response_model=RaffleResponse)
async def get_archived_raffle(raffle_id: str):
    raffle = await db.raffles.find_one(
        {"id": raffle_id, "status": {"$in": [RaffleStatus.INACTIVE, RaffleStatus.FINISHED]}},
        {"_id": 0}
    )
    if not raffle:
        raise HTTPException(status_code=404, detail="Rifa archivada no encontrada")
    return raffle

# ============ WHATSAPP CLICK-TO-CHAT ============
class BusinessConfigUpdate(BaseModel):
    business_whatsapp: Optional[str] = None
    business_name: Optional[str] = None

@api_router.get("/config/business")
async def get_business_config():
    config = await db.business_config.find_one({"type": "main"}, {"_id": 0})
    if not config:
        return {"business_whatsapp": "", "business_name": "SuerteApp"}
    return config

@api_router.put("/config/business")
async def update_business_config(config_data: BusinessConfigUpdate, admin: dict = Depends(require_admin)):
    update_dict = {k: v for k, v in config_data.model_dump().items() if v is not None}
    update_dict["type"] = "main"
    
    await db.business_config.update_one(
        {"type": "main"},
        {"$set": update_dict},
        upsert=True
    )
    return {"message": "Configuración actualizada"}

@api_router.get("/whatsapp/link")
async def generate_whatsapp_link(
    raffle_id: str,
    slot_numbers: str,
    user: dict = Depends(get_current_user)
):
    """Generate WhatsApp click-to-chat link with purchase details"""
    raffle = await db.raffles.find_one({"id": raffle_id}, {"_id": 0})
    if not raffle:
        raise HTTPException(status_code=404, detail="Rifa no encontrada")
    
    config = await db.business_config.find_one({"type": "main"}, {"_id": 0})
    business_whatsapp = config.get("business_whatsapp", "") if config else ""
    business_name = config.get("business_name", "SuerteApp") if config else "SuerteApp"
    
    if not business_whatsapp:
        raise HTTPException(status_code=400, detail="Número de WhatsApp del negocio no configurado")
    
    # Clean phone number
    clean_number = business_whatsapp.replace("+", "").replace(" ", "").replace("-", "")
    
    # Parse slot numbers
    slots = slot_numbers.split(",")
    total_amount = len(slots) * raffle["slot_price"]
    
    # Build message
    message = f"""🎰 *RESERVA DE NÚMEROS - {business_name}*

📋 *Rifa:* {raffle['title']}
🔢 *Números:* {slot_numbers}
💰 *Total:* ${total_amount:.2f}

👤 *Datos del comprador:*
• Nombre: {user['full_name']}
• Cédula: {user['cedula']}
• WhatsApp: {user['whatsapp']}

✅ Por favor confirma mi reserva y envíame los datos para el pago."""

    # URL encode the message
    import urllib.parse
    encoded_message = urllib.parse.quote(message)
    
    whatsapp_url = f"https://wa.me/{clean_number}?text={encoded_message}"
    
    return {
        "whatsapp_url": whatsapp_url,
        "business_name": business_name,
        "message": message
    }

# ============ STATS ============
@api_router.get("/stats")
async def get_stats(admin: dict = Depends(require_admin)):
    total_users = await db.users.count_documents({})
    total_raffles = await db.raffles.count_documents({})
    active_raffles = await db.raffles.count_documents({"status": RaffleStatus.ACTIVE})
    total_transactions = await db.payment_transactions.count_documents({"status": PaymentStatus.COMPLETED})
    
    pipeline = [
        {"$match": {"status": PaymentStatus.COMPLETED}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    revenue_result = await db.payment_transactions.aggregate(pipeline).to_list(1)
    total_revenue = revenue_result[0]["total"] if revenue_result else 0
    
    # Sponsor stats
    sponsor_pipeline = [
        {"$group": {"_id": None, "total": {"$sum": "$commission_amount"}}}
    ]
    sponsor_result = await db.sponsor_earnings.aggregate(sponsor_pipeline).to_list(1)
    total_sponsor_commissions = sponsor_result[0]["total"] if sponsor_result else 0
    
    return {
        "total_users": total_users,
        "total_raffles": total_raffles,
        "active_raffles": active_raffles,
        "completed_transactions": total_transactions,
        "total_revenue": total_revenue,
        "total_sponsor_commissions": total_sponsor_commissions
    }

# ============ SPONSOR PROGRAM ============
@api_router.get("/sponsor/my-code")
async def get_my_sponsor_code(user: dict = Depends(get_current_user)):
    """Get user's sponsor code and stats"""
    # Get total earnings
    pipeline = [
        {"$match": {"sponsor_id": user["id"]}},
        {"$group": {
            "_id": None,
            "total_earnings": {"$sum": "$commission_amount"},
            "pending_earnings": {
                "$sum": {"$cond": [{"$eq": ["$status", "pending"]}, "$commission_amount", 0]}
            },
            "paid_earnings": {
                "$sum": {"$cond": [{"$eq": ["$status", "paid"]}, "$commission_amount", 0]}
            },
            "total_referrals": {"$sum": 1}
        }}
    ]
    stats_result = await db.sponsor_earnings.aggregate(pipeline).to_list(1)
    stats = stats_result[0] if stats_result else {
        "total_earnings": 0,
        "pending_earnings": 0,
        "paid_earnings": 0,
        "total_referrals": 0
    }
    
    # Count referred users
    referred_users = await db.users.count_documents({"referred_by": user["id"]})
    
    return {
        "sponsor_code": user.get("sponsor_code", ""),
        "total_earnings": stats.get("total_earnings", 0),
        "pending_earnings": stats.get("pending_earnings", 0),
        "paid_earnings": stats.get("paid_earnings", 0),
        "total_sales": stats.get("total_referrals", 0),
        "referred_users": referred_users,
        "commission_rate": SPONSOR_COMMISSION_RATE * 100
    }

@api_router.get("/sponsor/my-earnings")
async def get_my_earnings(user: dict = Depends(get_current_user)):
    """Get detailed earnings history"""
    earnings = await db.sponsor_earnings.find(
        {"sponsor_id": user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return earnings

@api_router.get("/sponsor/my-weekly-summary")
async def get_my_weekly_summary(user: dict = Depends(get_current_user)):
    """Get current week summary for the logged-in sponsor"""
    current_week = datetime.now(timezone.utc).strftime("%Y-W%W")
    
    # Get this week's earnings
    week_earnings = await db.sponsor_earnings.find(
        {"sponsor_id": user["id"], "week_year": current_week},
        {"_id": 0}
    ).to_list(100)
    
    # Get all pending earnings (not just this week)
    all_pending = await db.sponsor_earnings.find(
        {"sponsor_id": user["id"], "status": "pending"},
        {"_id": 0}
    ).to_list(100)
    
    total_week = sum(e["commission_amount"] for e in week_earnings)
    total_pending = sum(e["commission_amount"] for e in all_pending)
    
    # Check if there's already a pending payment request
    existing_request = await db.payment_requests.find_one(
        {"sponsor_id": user["id"], "status": "pending"},
        {"_id": 0}
    )
    
    return {
        "week": current_week,
        "week_earnings": total_week,
        "week_sales": len(week_earnings),
        "total_pending": total_pending,
        "pending_sales": len(all_pending),
        "can_request_payment": total_pending > 0 and not existing_request,
        "has_pending_request": existing_request is not None,
        "earnings_details": week_earnings
    }

@api_router.post("/sponsor/request-payment")
async def request_payment(request: PaymentRequestCreate, user: dict = Depends(get_current_user)):
    """Request payment for pending sponsor earnings"""
    # Get pending earnings
    pending_earnings = await db.sponsor_earnings.find(
        {"sponsor_id": user["id"], "status": "pending"},
        {"_id": 0}
    ).to_list(100)
    
    total_pending = sum(e["commission_amount"] for e in pending_earnings)
    
    if total_pending <= 0:
        raise HTTPException(status_code=400, detail="No tienes ganancias pendientes")
    
    if request.amount > total_pending:
        raise HTTPException(status_code=400, detail=f"El monto solicitado excede tu saldo pendiente (${total_pending:.2f})")
    
    # Check for existing pending request
    existing = await db.payment_requests.find_one(
        {"sponsor_id": user["id"], "status": "pending"},
        {"_id": 0}
    )
    if existing:
        raise HTTPException(status_code=400, detail="Ya tienes una solicitud de pago pendiente")
    
    # Create payment request
    request_id = str(uuid.uuid4())
    payment_request = {
        "id": request_id,
        "sponsor_id": user["id"],
        "sponsor_name": user["full_name"],
        "sponsor_email": user["email"],
        "sponsor_whatsapp": user["whatsapp"],
        "bank_name": request.bank_name,
        "account_number": request.account_number,
        "account_holder": request.account_holder,
        "amount": request.amount,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "processed_at": None
    }
    await db.payment_requests.insert_one(payment_request)
    
    # Send email to admin
    admin_email = os.environ.get('ADMIN_EMAIL', '')
    if admin_email and resend.api_key:
        try:
            config = await db.business_config.find_one({"type": "main"}, {"_id": 0})
            business_name = config.get("business_name", "SuerteApp") if config else "SuerteApp"
            
            html_content = f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <h2 style="color: #003366;">💰 Nueva Solicitud de Pago - {business_name}</h2>
                <p>Un sponsor ha solicitado el retiro de sus ganancias:</p>
                
                <div style="background: #f5f5f5; padding: 20px; border-radius: 8px; margin: 20px 0;">
                    <h3 style="margin-top: 0;">Datos del Sponsor:</h3>
                    <p><strong>Nombre:</strong> {user['full_name']}</p>
                    <p><strong>Email:</strong> {user['email']}</p>
                    <p><strong>WhatsApp:</strong> {user['whatsapp']}</p>
                    <p><strong>Código Sponsor:</strong> {user.get('sponsor_code', 'N/A')}</p>
                </div>
                
                <div style="background: #e8f5e9; padding: 20px; border-radius: 8px; margin: 20px 0;">
                    <h3 style="margin-top: 0;">Datos Bancarios:</h3>
                    <p><strong>Banco:</strong> {request.bank_name}</p>
                    <p><strong>Número de Cuenta:</strong> {request.account_number}</p>
                    <p><strong>Titular:</strong> {request.account_holder}</p>
                    <p style="font-size: 24px; color: #2e7d32;"><strong>Monto: ${request.amount:.2f}</strong></p>
                </div>
                
                <p style="color: #666;">Por favor, procesa esta solicitud desde el panel de administración.</p>
            </div>
            """
            
            params = {
                "from": SENDER_EMAIL,
                "to": [admin_email],
                "subject": f"💰 Solicitud de Pago - {user['full_name']} - ${request.amount:.2f}",
                "html": html_content
            }
            
            await asyncio.to_thread(resend.Emails.send, params)
            logger.info(f"Payment request email sent for {user['full_name']}")
        except Exception as e:
            logger.error(f"Failed to send payment request email: {e}")
    
    return {
        "message": "Solicitud de pago enviada correctamente",
        "request_id": request_id,
        "amount": request.amount
    }

@api_router.get("/sponsor/my-payment-requests")
async def get_my_payment_requests(user: dict = Depends(get_current_user)):
    """Get user's payment request history"""
    requests = await db.payment_requests.find(
        {"sponsor_id": user["id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return requests

@api_router.get("/admin/payment-requests")
async def get_all_payment_requests(admin: dict = Depends(require_admin)):
    """Get all payment requests (admin only)"""
    requests = await db.payment_requests.find(
        {},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return requests

@api_router.put("/admin/payment-requests/{request_id}/approve")
async def approve_payment_request(request_id: str, admin: dict = Depends(require_admin)):
    """Approve a payment request"""
    request = await db.payment_requests.find_one({"id": request_id}, {"_id": 0})
    if not request:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    
    if request["status"] != "pending":
        raise HTTPException(status_code=400, detail="Esta solicitud ya fue procesada")
    
    # Update request status
    await db.payment_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "approved", "processed_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Mark sponsor earnings as paid (up to the requested amount)
    pending_earnings = await db.sponsor_earnings.find(
        {"sponsor_id": request["sponsor_id"], "status": "pending"},
        {"_id": 0}
    ).sort("created_at", 1).to_list(100)
    
    remaining_amount = request["amount"]
    for earning in pending_earnings:
        if remaining_amount <= 0:
            break
        if earning["commission_amount"] <= remaining_amount:
            await db.sponsor_earnings.update_one(
                {"id": earning["id"]},
                {"$set": {"status": "paid", "paid_at": datetime.now(timezone.utc).isoformat()}}
            )
            remaining_amount -= earning["commission_amount"]
    
    return {"message": "Solicitud aprobada"}

@api_router.put("/admin/payment-requests/{request_id}/reject")
async def reject_payment_request(request_id: str, admin: dict = Depends(require_admin)):
    """Reject a payment request"""
    result = await db.payment_requests.update_one(
        {"id": request_id, "status": "pending"},
        {"$set": {"status": "rejected", "processed_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada o ya procesada")
    return {"message": "Solicitud rechazada"}

@api_router.put("/admin/payment-requests/{request_id}/mark-paid")
async def mark_payment_request_paid(request_id: str, admin: dict = Depends(require_admin)):
    """Mark payment request as paid"""
    result = await db.payment_requests.update_one(
        {"id": request_id, "status": "approved"},
        {"$set": {"status": "paid", "paid_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada o no aprobada")
    return {"message": "Pago marcado como realizado"}

@api_router.get("/sponsor/weekly-summary/{user_id}")
async def get_weekly_summary(user_id: str, admin: dict = Depends(require_admin)):
    """Get weekly summary for a sponsor (admin only)"""
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Get current week
    current_week = datetime.now(timezone.utc).strftime("%Y-W%W")
    
    # Get this week's earnings
    earnings = await db.sponsor_earnings.find(
        {"sponsor_id": user_id, "week_year": current_week},
        {"_id": 0}
    ).to_list(100)
    
    total_week = sum(e["commission_amount"] for e in earnings)
    
    return {
        "user": user,
        "week": current_week,
        "earnings": earnings,
        "total_week_earnings": total_week
    }

@api_router.get("/sponsor/all-sponsors")
async def get_all_sponsors(admin: dict = Depends(require_admin)):
    """Get all sponsors with their stats (admin only)"""
    # Get all users with sponsor earnings
    pipeline = [
        {"$group": {
            "_id": "$sponsor_id",
            "total_earnings": {"$sum": "$commission_amount"},
            "pending_earnings": {
                "$sum": {"$cond": [{"$eq": ["$status", "pending"]}, "$commission_amount", 0]}
            },
            "total_sales": {"$sum": 1}
        }}
    ]
    earnings_by_sponsor = await db.sponsor_earnings.aggregate(pipeline).to_list(100)
    
    sponsors = []
    for earning in earnings_by_sponsor:
        user = await db.users.find_one({"id": earning["_id"]}, {"_id": 0, "password": 0})
        if user:
            sponsors.append({
                "user": user,
                "total_earnings": earning["total_earnings"],
                "pending_earnings": earning["pending_earnings"],
                "total_sales": earning["total_sales"]
            })
    
    return sponsors

@api_router.post("/sponsor/mark-paid/{earning_id}")
async def mark_earning_paid(earning_id: str, admin: dict = Depends(require_admin)):
    """Mark a sponsor earning as paid"""
    result = await db.sponsor_earnings.update_one(
        {"id": earning_id},
        {"$set": {"status": "paid", "paid_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Ganancia no encontrada")
    return {"message": "Marcado como pagado"}

@api_router.post("/sponsor/pay-all/{user_id}")
async def pay_all_pending(user_id: str, admin: dict = Depends(require_admin)):
    """Mark all pending earnings for a sponsor as paid"""
    result = await db.sponsor_earnings.update_many(
        {"sponsor_id": user_id, "status": "pending"},
        {"$set": {"status": "paid", "paid_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": f"{result.modified_count} ganancias marcadas como pagadas"}

@api_router.get("/sponsor/weekly-report")
async def get_weekly_report(admin: dict = Depends(require_admin)):
    """Get weekly report for all sponsors with pending earnings"""
    current_week = datetime.now(timezone.utc).strftime("%Y-W%W")
    
    # Get all pending earnings grouped by sponsor
    pipeline = [
        {"$match": {"status": "pending"}},
        {"$group": {
            "_id": "$sponsor_id",
            "total_pending": {"$sum": "$commission_amount"},
            "sales_count": {"$sum": 1},
            "earnings": {"$push": "$$ROOT"}
        }}
    ]
    results = await db.sponsor_earnings.aggregate(pipeline).to_list(100)
    
    report = []
    for item in results:
        user = await db.users.find_one({"id": item["_id"]}, {"_id": 0, "password": 0})
        if user:
            report.append({
                "sponsor": user,
                "total_pending": item["total_pending"],
                "sales_count": item["sales_count"],
                "whatsapp": user.get("whatsapp", ""),
                "earnings_details": item["earnings"]
            })
    
    return {"week": current_week, "sponsors": report}

@api_router.post("/sponsor/send-weekly-whatsapp")
async def send_weekly_whatsapp(admin: dict = Depends(require_admin)):
    """Generate WhatsApp messages for all sponsors with pending earnings"""
    current_week = datetime.now(timezone.utc).strftime("%Y-W%W")
    
    # Get config
    config = await db.business_config.find_one({"type": "main"}, {"_id": 0})
    business_name = config.get("business_name", "SuerteApp") if config else "SuerteApp"
    
    # Get all pending earnings grouped by sponsor
    pipeline = [
        {"$match": {"status": "pending"}},
        {"$group": {
            "_id": "$sponsor_id",
            "total_pending": {"$sum": "$commission_amount"},
            "sales_count": {"$sum": 1}
        }}
    ]
    results = await db.sponsor_earnings.aggregate(pipeline).to_list(100)
    
    messages = []
    for item in results:
        user = await db.users.find_one({"id": item["_id"]}, {"_id": 0})
        if user and user.get("whatsapp"):
            message = f"""🎉 *RESUMEN SEMANAL DE GANANCIAS - {business_name}*

¡Hola {user['full_name'].split()[0]}! 👋

📊 *Tu resumen de la semana:*
• Ventas realizadas: {item['sales_count']}
• Comisiones generadas: ${item['total_pending']:.2f}

💰 *Total a cobrar: ${item['total_pending']:.2f}*

Tu código de sponsor: *{user.get('sponsor_code', 'N/A')}*

¡Sigue compartiendo tu código para ganar más! 🚀

Gracias por ser parte de {business_name}. 💙"""
            
            import urllib.parse
            clean_number = user["whatsapp"].replace("+", "").replace(" ", "").replace("-", "")
            whatsapp_url = f"https://wa.me/{clean_number}?text={urllib.parse.quote(message)}"
            
            messages.append({
                "sponsor_id": user["id"],


# ============ PASSWORD RECOVERY ============
class ForgotPasswordRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str = Field(..., min_length=6)

def create_reset_token(email: str) -> str:
    """Create a password reset token valid for 1 hour"""
    payload = {
        "email": email,
        "type": "password_reset",
        "exp": datetime.now(timezone.utc) + timedelta(hours=1)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    """Send password reset email"""
    user = await db.users.find_one({"email": request.email}, {"_id": 0})
    if not user:
        # Don't reveal if email exists for security
        return {"message": "Si el email existe, recibirás un enlace de recuperación"}
    
    # Create reset token
    reset_token = create_reset_token(request.email)
    
    # Get business name from config
    config = await db.business_config.find_one({"type": "main"}, {"_id": 0})
    business_name = config.get("business_name", "SuerteApp") if config else "SuerteApp"
    
    # Get frontend URL from env
    frontend_url = os.environ.get('FRONTEND_URL', 'https://ventana-proyecto.preview.emergentagent.com')
    reset_link = f"{frontend_url}/reset-password?token={reset_token}"
    
    # Send email
    if resend.api_key:
        try:
            html_content = f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #003366;">🔐 Recuperación de Contraseña - {business_name}</h2>
                <p>Hola <strong>{user['full_name']}</strong>,</p>
                <p>Recibimos una solicitud para restablecer tu contraseña.</p>
                
                <div style="background: #f5f5f5; padding: 20px; border-radius: 8px; margin: 20px 0; text-align: center;">
                    <a href="{reset_link}" 
                       style="display: inline-block; background: #003366; color: white; 
                              padding: 12px 30px; text-decoration: none; border-radius: 5px; 
                              font-weight: bold;">
                        Restablecer Contraseña
                    </a>
                </div>
                
                <p style="color: #666; font-size: 14px;">
                    Este enlace es válido por <strong>1 hora</strong>.
                </p>
                <p style="color: #666; font-size: 14px;">
                    Si no solicitaste este cambio, puedes ignorar este email de forma segura.
                </p>
                
                <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
                <p style="color: #999; font-size: 12px;">
                    {business_name} - Plataforma de Rifas
                </p>
            </div>
            """
            
            params = {
                "from": SENDER_EMAIL,
                "to": [request.email],
                "subject": f"Recupera tu contraseña - {business_name}",
                "html": html_content
            }
            
            await asyncio.to_thread(resend.Emails.send, params)
            logger.info(f"Password reset email sent to {request.email}")
        except Exception as e:
            logger.error(f"Failed to send reset email: {e}")
            raise HTTPException(status_code=500, detail="Error al enviar el email")
    
    return {"message": "Si el email existe, recibirás un enlace de recuperación"}

@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    """Reset password using token"""
    try:
        payload = jwt.decode(request.token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "password_reset":
            raise HTTPException(status_code=400, detail="Token inválido")
        
        email = payload.get("email")
        user = await db.users.find_one({"email": email}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
        # Update password
        new_password_hash = hash_password(request.new_password)
        await db.users.update_one(
            {"email": email},
            {"$set": {"password": new_password_hash}}
        )
        
        logger.info(f"Password reset successful for {email}")
        return {"message": "Contraseña actualizada exitosamente"}
        
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=400, detail="El token ha expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=400, detail="Token inválido")

# ============ USER EDITING (ADMIN) ============
class UserUpdateRequest(BaseModel):
    full_name: Optional[str] = None
    email: Optional[EmailStr] = None
    whatsapp: Optional[str] = None
    cedula: Optional[str] = None
    role: Optional[UserRole] = None

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: str, user_data: UserUpdateRequest, admin: dict = Depends(require_admin)):
    """Update user information (admin only)"""
    update_dict = {k: v for k, v in user_data.model_dump().items() if v is not None}
    
    if not update_dict:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    
    # Check if email or cedula already exists (if being updated)
    if "email" in update_dict or "cedula" in update_dict:
        query_conditions = []
        if "email" in update_dict:
            query_conditions.append({"email": update_dict["email"]})
        if "cedula" in update_dict:
            query_conditions.append({"cedula": update_dict["cedula"]})
        
        existing = await db.users.find_one({
            "$and": [
                {"id": {"$ne": user_id}},
                {"$or": query_conditions}
            ]
        })
        if existing:
            raise HTTPException(status_code=400, detail="Email o cédula ya están en uso")
    
    result = await db.users.find_one_and_update(
        {"id": user_id},
        {"$set": update_dict},
        return_document=True
    )
    
    if not result:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    return {k: v for k, v in result.items() if k not in ["password", "_id"]}

# ============ REPORTS GENERATION ============
class ReportType(str, Enum):
    EXCEL = "excel"
    PDF = "pdf"

class ReportHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    report_type: ReportType
    generated_by: str
    generated_by_name: str
    filename: str
    created_at: str

async def get_report_data():
    """Get all data needed for reports"""
    # Get stats
    total_users = await db.users.count_documents({})
    total_raffles = await db.raffles.count_documents({})
    active_raffles = await db.raffles.count_documents({"status": RaffleStatus.ACTIVE})
    
    # Get completed transactions
    completed_txns = await db.payment_transactions.find(
        {"status": PaymentStatus.COMPLETED},
        {"_id": 0}
    ).to_list(1000)
    
    total_revenue = sum(t["amount"] for t in completed_txns)
    
    # Group by week
    week_revenue = {}
    for txn in completed_txns:
        created = datetime.fromisoformat(txn["created_at"])
        week_key = created.strftime("%Y-W%W")
        week_revenue[week_key] = week_revenue.get(week_key, 0) + txn["amount"]
    
    # Group by month
    month_revenue = {}
    for txn in completed_txns:
        created = datetime.fromisoformat(txn["created_at"])
        month_key = created.strftime("%Y-%m")
        month_revenue[month_key] = month_revenue.get(month_key, 0) + txn["amount"]
    
    # Group by sponsor
    sponsor_earnings = await db.sponsor_earnings.find({}, {"_id": 0}).to_list(1000)
    sponsor_totals = {}
    for earning in sponsor_earnings:
        sponsor_id = earning["sponsor_id"]
        if sponsor_id not in sponsor_totals:
            user = await db.users.find_one({"id": sponsor_id}, {"_id": 0, "password": 0})
            sponsor_totals[sponsor_id] = {
                "name": user["full_name"] if user else "Desconocido",
                "total": 0,
                "pending": 0,
                "paid": 0
            }
        sponsor_totals[sponsor_id]["total"] += earning["commission_amount"]
        if earning["status"] == "pending":
            sponsor_totals[sponsor_id]["pending"] += earning["commission_amount"]
        elif earning["status"] == "paid":
            sponsor_totals[sponsor_id]["paid"] += earning["commission_amount"]
    
    # Group by user (buyers)
    user_purchases = {}
    for txn in completed_txns:
        user_id = txn["user_id"]
        if user_id not in user_purchases:
            user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
            user_purchases[user_id] = {
                "name": user["full_name"] if user else "Desconocido",
                "total_spent": 0,
                "transactions": 0
            }
        user_purchases[user_id]["total_spent"] += txn["amount"]
        user_purchases[user_id]["transactions"] += 1
    
    # Group by raffle
    raffle_sales = {}
    for txn in completed_txns:
        raffle_id = txn["raffle_id"]
        if raffle_id not in raffle_sales:
            raffle = await db.raffles.find_one({"id": raffle_id}, {"_id": 0})
            raffle_sales[raffle_id] = {
                "title": raffle["title"] if raffle else "Desconocida",
                "revenue": 0,
                "slots_sold": 0
            }
        raffle_sales[raffle_id]["revenue"] += txn["amount"]
        raffle_sales[raffle_id]["slots_sold"] += len(txn["slot_numbers"])
    
    return {
        "total_users": total_users,
        "total_raffles": total_raffles,
        "active_raffles": active_raffles,
        "total_transactions": len(completed_txns),
        "total_revenue": total_revenue,
        "week_revenue": week_revenue,
        "month_revenue": month_revenue,
        "sponsor_totals": sponsor_totals,
        "user_purchases": user_purchases,
        "raffle_sales": raffle_sales
    }

@api_router.get("/reports/excel")
async def generate_excel_report(admin: dict = Depends(require_admin)):
    """Generate Excel report with all metrics"""
    data = await get_report_data()
    
    # Create workbook
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Resumen General"
    
    # Header styling
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    ws_summary['A1'] = "REPORTE GENERAL - SUERTEAPP"
    ws_summary['A1'].font = Font(size=16, bold=True, color="003366")
    ws_summary.merge_cells('A1:B1')
    
    ws_summary['A3'] = "Métrica"
    ws_summary['B3'] = "Valor"
    ws_summary['A3'].fill = header_fill
    ws_summary['A3'].font = header_font
    ws_summary['B3'].fill = header_fill
    ws_summary['B3'].font = header_font
    
    summary_data = [
        ["Total Usuarios", data["total_users"]],
        ["Total Rifas", data["total_raffles"]],
        ["Rifas Activas", data["active_raffles"]],
        ["Transacciones Completadas", data["total_transactions"]],
        ["Ingresos Totales", f"${data['total_revenue']:.2f}"]
    ]
    
    for idx, row in enumerate(summary_data, start=4):
        ws_summary[f'A{idx}'] = row[0]
        ws_summary[f'B{idx}'] = row[1]
    
    # Weekly revenue sheet
    ws_week = wb.create_sheet("Ingresos por Semana")
    ws_week['A1'] = "Semana"
    ws_week['B1'] = "Ingresos"
    ws_week['A1'].fill = header_fill
    ws_week['A1'].font = header_font
    ws_week['B1'].fill = header_fill
    ws_week['B1'].font = header_font
    
    for idx, (week, revenue) in enumerate(sorted(data["week_revenue"].items()), start=2):
        ws_week[f'A{idx}'] = week
        ws_week[f'B{idx}'] = f"${revenue:.2f}"
    
    # Monthly revenue sheet
    ws_month = wb.create_sheet("Ingresos por Mes")
    ws_month['A1'] = "Mes"
    ws_month['B1'] = "Ingresos"
    ws_month['A1'].fill = header_fill
    ws_month['A1'].font = header_font
    ws_month['B1'].fill = header_fill
    ws_month['B1'].font = header_font
    
    for idx, (month, revenue) in enumerate(sorted(data["month_revenue"].items()), start=2):
        ws_month[f'A{idx}'] = month
        ws_month[f'B{idx}'] = f"${revenue:.2f}"
    
    # Sponsor earnings sheet
    ws_sponsor = wb.create_sheet("Ganancias por Sponsor")
    ws_sponsor['A1'] = "Sponsor"
    ws_sponsor['B1'] = "Total"
    ws_sponsor['C1'] = "Pendiente"
    ws_sponsor['D1'] = "Pagado"
    for col in ['A1', 'B1', 'C1', 'D1']:
        ws_sponsor[col].fill = header_fill
        ws_sponsor[col].font = header_font
    
    for idx, (sponsor_id, info) in enumerate(data["sponsor_totals"].items(), start=2):
        ws_sponsor[f'A{idx}'] = info["name"]
        ws_sponsor[f'B{idx}'] = f"${info['total']:.2f}"
        ws_sponsor[f'C{idx}'] = f"${info['pending']:.2f}"
        ws_sponsor[f'D{idx}'] = f"${info['paid']:.2f}"
    
    # User purchases sheet
    ws_users = wb.create_sheet("Compras por Usuario")
    ws_users['A1'] = "Usuario"
    ws_users['B1'] = "Total Gastado"
    ws_users['C1'] = "Transacciones"
    for col in ['A1', 'B1', 'C1']:
        ws_users[col].fill = header_fill
        ws_users[col].font = header_font
    
    for idx, (user_id, info) in enumerate(data["user_purchases"].items(), start=2):
        ws_users[f'A{idx}'] = info["name"]
        ws_users[f'B{idx}'] = f"${info['total_spent']:.2f}"
        ws_users[f'C{idx}'] = info["transactions"]
    
    # Raffle sales sheet
    ws_raffles = wb.create_sheet("Ventas por Rifa")
    ws_raffles['A1'] = "Rifa"
    ws_raffles['B1'] = "Ingresos"
    ws_raffles['C1'] = "Casillas Vendidas"
    for col in ['A1', 'B1', 'C1']:
        ws_raffles[col].fill = header_fill
        ws_raffles[col].font = header_font
    
    for idx, (raffle_id, info) in enumerate(data["raffle_sales"].items(), start=2):
        ws_raffles[f'A{idx}'] = info["title"]
        ws_raffles[f'B{idx}'] = f"${info['revenue']:.2f}"
        ws_raffles[f'C{idx}'] = info["slots_sold"]
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Save report history
    filename = f"reporte_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    report_doc = {
        "id": str(uuid.uuid4()),
        "report_type": ReportType.EXCEL,
        "generated_by": admin["id"],
        "generated_by_name": admin["full_name"],
        "filename": filename,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.report_history.insert_one(report_doc)
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/pdf")
async def generate_pdf_report(admin: dict = Depends(require_admin)):
    """Generate PDF report with all metrics"""
    data = await get_report_data()
    
    # Create PDF
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("REPORTE GENERAL - SUERTEAPP", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary table
    summary_data = [
        ['Métrica', 'Valor'],
        ['Total Usuarios', str(data["total_users"])],
        ['Total Rifas', str(data["total_raffles"])],
        ['Rifas Activas', str(data["active_raffles"])],
        ['Transacciones Completadas', str(data["total_transactions"])],
        ['Ingresos Totales', f"${data['total_revenue']:.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Weekly revenue
    elements.append(Paragraph("Ingresos por Semana", styles['Heading2']))
    week_data = [['Semana', 'Ingresos']]
    for week, revenue in sorted(data["week_revenue"].items())[:10]:  # Last 10 weeks
        week_data.append([week, f"${revenue:.2f}"])
    
    week_table = Table(week_data, colWidths=[2.5*inch, 2*inch])
    week_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(week_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Top sponsors
    elements.append(Paragraph("Top 5 Sponsors", styles['Heading2']))
    sponsor_data = [['Sponsor', 'Total', 'Pendiente', 'Pagado']]
    sorted_sponsors = sorted(data["sponsor_totals"].items(), 
                            key=lambda x: x[1]["total"], reverse=True)[:5]
    for sponsor_id, info in sorted_sponsors:
        sponsor_data.append([
            info["name"][:30],
            f"${info['total']:.2f}",
            f"${info['pending']:.2f}",
            f"${info['paid']:.2f}"
        ])
    
    sponsor_table = Table(sponsor_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    sponsor_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(sponsor_table)
    
    # Build PDF
    doc.build(elements)
    pdf_buffer.seek(0)
    
    # Save report history
    filename = f"reporte_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.pdf"
    report_doc = {
        "id": str(uuid.uuid4()),
        "report_type": ReportType.PDF,
        "generated_by": admin["id"],
        "generated_by_name": admin["full_name"],
        "filename": filename,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.report_history.insert_one(report_doc)
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/history", response_model=List[ReportHistory])
async def get_report_history(admin: dict = Depends(require_admin)):
    """Get report generation history"""
    reports = await db.report_history.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return reports

# ============ FRONTEND CONTENT CMS ============
class FrontendContentUpdate(BaseModel):
    hero_title: Optional[str] = None
    hero_subtitle: Optional[str] = None
    hero_button_text: Optional[str] = None
    hero_image_url: Optional[str] = None
    logo_url: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    feature1_title: Optional[str] = None
    feature1_description: Optional[str] = None
    feature2_title: Optional[str] = None
    feature2_description: Optional[str] = None
    feature3_title: Optional[str] = None
    feature3_description: Optional[str] = None
    feature4_title: Optional[str] = None
    feature4_description: Optional[str] = None

class FrontendContentResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    hero_title: str
    hero_subtitle: str
    hero_button_text: str
    hero_image_url: str
    logo_url: str
    primary_color: str
    secondary_color: str
    feature1_title: str
    feature1_description: str
    feature2_title: str
    feature2_description: str
    feature3_title: str
    feature3_description: str
    feature4_title: str
    feature4_description: str
    updated_at: str

@api_router.get("/frontend/content", response_model=FrontendContentResponse)
async def get_frontend_content():
    """Get current frontend content configuration"""
    content = await db.frontend_content.find_one({"type": "main"}, {"_id": 0, "type": 0})
    
    if not content:
        # Return default content
        default_content = {
            "hero_title": "Tu Suerte Comienza Aquí",
            "hero_subtitle": "Participa en rifas emocionantes con premios increíbles. Elige tus números de la suerte y gana desde tu móvil.",
            "hero_button_text": "Participar Ahora",
            "hero_image_url": "https://images.pexels.com/photos/6612233/pexels-photo-6612233.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940",
            "logo_url": "https://avatars.githubusercontent.com/in/1201222?s=120&u=2686cf91179bbafbc7a71bfbc43004cf9ae1acea&v=4",
            "primary_color": "#003366",
            "secondary_color": "#28a745",
            "feature1_title": "Grandes Premios",
            "feature1_description": "Participa por increíbles premios cada semana",
            "feature2_title": "100% Seguro",
            "feature2_description": "Transacciones protegidas y transparentes",
            "feature3_title": "Desde tu Móvil",
            "feature3_description": "Compra tus números desde cualquier lugar",
            "feature4_title": "Fácil y Rápido",
            "feature4_description": "Elige tus números y paga en segundos",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        # Save default
        await db.frontend_content.insert_one({**default_content, "type": "main"})
        return default_content
    
    return content

@api_router.put("/frontend/content")
async def update_frontend_content(content_data: FrontendContentUpdate, admin: dict = Depends(require_admin)):
    """Update frontend content (admin only)"""
    update_dict = {k: v for k, v in content_data.model_dump().items() if v is not None}
    
    if not update_dict:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.frontend_content.update_one(
        {"type": "main"},
        {"$set": update_dict},
        upsert=True
    )
    
    return {"message": "Contenido actualizado exitosamente"}

                "sponsor_name": user["full_name"],
                "whatsapp": user["whatsapp"],
                "total_pending": item["total_pending"],
                "sales_count": item["sales_count"],
                "whatsapp_url": whatsapp_url,
                "message": message
            })
    
    return {"week": current_week, "messages": messages}

# Include router and middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
